#!/usr/bin/env python3
"""Importe les anniversaires depuis le fichier Excel vers data/birthdays.json."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl

MONTHS_FR = [
    "Janvier",
    "Février",
    "Mars",
    "Avril",
    "Mai",
    "Juin",
    "Juillet",
    "Août",
    "Septembre",
    "Octobre",
    "Novembre",
    "Décembre",
]
MONTH_MAP = {name: index + 1 for index, name in enumerate(MONTHS_FR)}


def parse_birthdays(xlsx_path: Path) -> list[dict[str, int | str]]:
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook.active

    header_row = list(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    month_columns: dict[int, int] = {}
    for column_index, value in enumerate(header_row):
        if value and str(value).strip() in MONTH_MAP:
            month_columns[column_index] = MONTH_MAP[str(value).strip()]

    birthdays: list[dict[str, int | str]] = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        for day_column, month_number in month_columns.items():
            day_value = row[day_column] if day_column < len(row) else None
            name_value = row[day_column + 1] if day_column + 1 < len(row) else None
            if not day_value or not isinstance(day_value, (int, float)):
                continue
            if not name_value:
                continue

            name = str(name_value).strip()
            if not name or name.replace(".", "").isdigit():
                continue

            birthdays.append(
                {
                    "month": month_number,
                    "day": int(day_value),
                    "name": name,
                }
            )

    birthdays.sort(key=lambda entry: (entry["month"], entry["day"]))
    return birthdays


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("data/source/Anniversaires.xlsx"),
        help="Chemin du fichier Excel source",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/birthdays.json"),
        help="Chemin du fichier JSON de sortie",
    )
    args = parser.parse_args()

    birthdays = parse_birthdays(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(birthdays, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"{len(birthdays)} anniversaires exportés vers {args.output}")


if __name__ == "__main__":
    main()
